import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("tx_scanner.writer")


class ExcelWriter:
    def __init__(self, excel_path: str, sheet_name: str, columns: list[str]):
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.columns = columns

    def initialize(self):
        """Create new Excel file with headers if it doesn't exist."""
        if not self.excel_path.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.sheet_name

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            for col_idx, col_name in enumerate(self.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            self._auto_width(ws)
            wb.save(str(self.excel_path))
            logger.info("Created Excel file: %s", self.excel_path)

    def append_rows(self, rows: list[dict]):
        """Append transaction rows to the Excel file."""
        if not rows:
            return

        if not self.excel_path.exists():
            self.initialize()

        wb = openpyxl.load_workbook(str(self.excel_path))

        if self.sheet_name not in wb.sheetnames:
            ws = wb.active
            ws.title = self.sheet_name
        else:
            ws = wb[self.sheet_name]

        for row in rows:
            ws.append([
                row.get("交易方", "未知"),
                row.get("交易时间", "未知"),
                row.get("支付方式", "未知"),
                row.get("账号", "未知"),
                row.get("交易金额", 0.0),
                row.get("对应图片", ""),
            ])

        self._auto_width(ws)
        wb.save(str(self.excel_path))
        logger.info("Appended %d row(s) to %s", len(rows), self.excel_path.name)

    def _auto_width(self, ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
